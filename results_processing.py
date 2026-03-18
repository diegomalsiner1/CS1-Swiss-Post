import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
# Add charts directly into the Excel workbook
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, BarChart

def compute_financial_summary(input_dict, solution_summary):
    lifetime = int(input_dict["parameters"].get("lifetime", 20))
    discount_rate = float(input_dict["parameters"].get("interest_rate", 0.06))
    invest_cost_per_kwh = float(input_dict["parameters"].get("Battery_invest_cost", 0.0))
    battery_capacity = float(solution_summary.get("battery_capacity_kwh", 0.0))

    investment = -battery_capacity * invest_cost_per_kwh
    # operational savings relative to no battery import cost
    no_battery_import_cost = float(solution_summary.get("no_battery_import_cost", np.nan))
    battery_import_cost = float(solution_summary.get("import_cost", np.nan))
    annual_savings = no_battery_import_cost - battery_import_cost

    # Make yearly cashflow (year0 = -investment, then annual_savings for each year)
    cashflows = [investment] + [annual_savings] * lifetime

    # Discount factors
    years = list(range(0, lifetime + 1))
    discount_factors = [(1 + discount_rate) ** (-y) for y in years]
    discounted_cashflows = [cf * df for cf, df in zip(cashflows, discount_factors)]
    npv = sum(discounted_cashflows)

    irr = np.nan
    try:
        irr = np.irr(cashflows)
    except Exception:
        irr = np.nan

    # Payback period (undiscounted)
    cumulative = 0.0
    payback = np.nan
    for year, cf in enumerate(cashflows):
        cumulative += cf
        if cumulative >= 0 and year > 0:
            payback = year
            break

    annual_financials_df = pd.DataFrame({
        "year": years,
        "cashflow": cashflows,
        "discount_factor": discount_factors,
        "discounted_cashflow": discounted_cashflows,
    })

    financial_summary = {
        "investment_cost": -investment,
        "annual_savings": annual_savings,
        "npv": npv,
        "irr": irr,
        "payback_years": payback,
        "annual_financials_df": annual_financials_df,
    }
    return financial_summary


def export_results_excel(run_dir, solution_summary, battery_soc, timestamps=None, input_dict=None, pv_flow=None, grid_flow=None, total_load=None):
    """Export KPI summary and SOC chart data to a single Excel file and PNG chart."""
    run_dir = Path(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)

    # KPI sheet
    selected_keys = [
        "objective_total_cost",
        "battery_capacity_kwh",
        "opex",
        "import_cost",
        "fixed_om_cost",
        "annualized_battery_cost",
        "no_battery_import_cost",
        "no_battery_total_cost",
        "npv",
        "irr",
        "payback_years",
    ]

    kpi_data = {k: solution_summary.get(k, np.nan) for k in selected_keys}
    kpi_df = pd.DataFrame.from_dict(kpi_data, orient="index", columns=["value"]).reset_index()
    kpi_df.columns = ["metric", "value"]

    # SOC time series
    if timestamps is not None:
        ts = pd.to_datetime(timestamps, errors="coerce")
        if ts.isna().any() or len(ts) != len(battery_soc):
            n = min(len(ts), len(battery_soc))
            if n <= 0:
                soc_df = pd.DataFrame({"timestamp": np.arange(len(battery_soc)), "battery_soc": battery_soc})
            else:
                soc_df = pd.DataFrame({"timestamp": ts[:n].tolist(), "battery_soc": battery_soc[:n]})
        else:
            soc_df = pd.DataFrame({"timestamp": ts.tolist(), "battery_soc": battery_soc})
    else:
        soc_df = pd.DataFrame({"timestamp": np.arange(len(battery_soc)), "battery_soc": battery_soc})

    # Save Excel with KPI + SOC sheet
    out_excel = run_dir / "results_summary.xlsx"
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        kpi_df.to_excel(writer, sheet_name="KPIs", index=False)
        soc_df.to_excel(writer, sheet_name="Battery_SOC", index=False)



    wb = load_workbook(out_excel)
    ws_soc = wb["Battery_SOC"]


    # Plot 4-in-one for Jan1/Apr1/Jul1/Oct1 one-week windows
    fig, axes = plt.subplots(2, 2, figsize=(11, 8), constrained_layout=True)
    axes = axes.flatten()
    month_days = [(1, 1), (4, 1), (7, 1), (10, 1)]

    if "timestamp" in soc_df.columns and np.issubdtype(soc_df["timestamp"].dtype, np.datetime64):
        if len(soc_df) > 0:
            year = soc_df["timestamp"].dt.year.iloc[0]
        else:
            year = datetime.now().year

        for i, (m, d) in enumerate(month_days):
            target_date = pd.Timestamp(year=year, month=m, day=d)
            window_start = target_date
            window_end = target_date + pd.Timedelta(days=7)
            window = soc_df[(soc_df["timestamp"] >= window_start) & (soc_df["timestamp"] < window_end)]
            ax = axes[i]
            if len(window) == 0:
                ax.text(0.5, 0.5, f"No data for {target_date.date()}", ha="center", va="center")
            else:
                ax.plot(window["timestamp"], window["battery_soc"], marker="o", markersize=3, linestyle="-", label=str(target_date.date()))
                ax.set_xlabel("Time")
                ax.set_ylabel("Battery SOC")
                ax.grid(True)
                ax.legend()
            ax.set_title(f"Week from {target_date.date()}")
    else:
        for i in range(4):
            ax = axes[i]
            ax.plot(soc_df["timestamp"], soc_df["battery_soc"], label="SOC")
            ax.set_title("SOC series")
            ax.set_xlabel("index")
            ax.set_ylabel("Battery SOC")
            ax.grid(True)

    fig.suptitle("Battery SOC 4-in-one week windows")
    chart_path = run_dir / "battery_soc_4in1.png"
    fig.savefig(chart_path)
    plt.close(fig)

    # 4-week power-flow comparison plot (PV, Grid, Total Load)
    if timestamps is not None and pv_flow is not None and grid_flow is not None and total_load is not None:
        ts2 = pd.to_datetime(timestamps, errors="coerce")
        if ts2.isna().any() or len(ts2) != len(pv_flow) or len(ts2) != len(grid_flow) or len(ts2) != len(total_load):
            n = min(len(ts2), len(pv_flow), len(grid_flow), len(total_load))
            ts2 = ts2[:n]
            pv_flow_plot = np.array(pv_flow)[:n]
            grid_flow_plot = np.array(grid_flow)[:n]
            load_plot = np.array(total_load)[:n]
        else:
            pv_flow_plot = np.array(pv_flow)
            grid_flow_plot = np.array(grid_flow)
            load_plot = np.array(total_load)

        fig2, axes2 = plt.subplots(2, 2, figsize=(12, 10), constrained_layout=True)
        axes2 = axes2.flatten()
        month_days = [(1, 1), (4, 1), (7, 1), (10, 1)]
        for i, (m, d) in enumerate(month_days):
            year0 = ts2.year[0] if len(ts2) > 0 else datetime.now().year
            target_date = pd.Timestamp(year=year0, month=m, day=d)
            week_data = (ts2 >= target_date) & (ts2 < target_date + pd.Timedelta(days=7))
            ax = axes2[i]
            if week_data.sum() == 0:
                ax.text(0.5, 0.5, f"No data for {target_date.date()}", ha="center", va="center")
            else:
                ax.plot(ts2[week_data], pv_flow_plot[week_data], label="PV flow")
                ax.plot(ts2[week_data], grid_flow_plot[week_data], label="Grid flow")
                ax.plot(ts2[week_data], load_plot[week_data], label="Total load")
                ax.set_xlabel("Time")
                ax.set_ylabel("Power (kW)")
                ax.grid(True)
                ax.legend(fontsize="small")
            ax.set_title(f"Week from {target_date.date()}")
        fig2.suptitle("4-week power flow comparison (PV, Grid, Total Load)")
        flow_chart_path = run_dir / "power_flow_4week.png"
        fig2.savefig(flow_chart_path)
        plt.close(fig2)
    else:
        flow_chart_path = None

    financial_results = {}
    if input_dict is not None:
        financial_results = compute_financial_summary(input_dict, solution_summary)
        solution_summary["npv"] = financial_results["npv"]
        solution_summary["irr"] = financial_results["irr"]
        solution_summary["payback_years"] = financial_results["payback_years"]
        fin_df = financial_results["annual_financials_df"]
        with pd.ExcelWriter(out_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            fin_df.to_excel(writer, sheet_name="Financial_Cashflows", index=False)

        # embed finance chart
        wb = load_workbook(out_excel)
        ws_fin = wb["Financial_Cashflows"]
        bar = BarChart()
        bar.title = "Cashflows and Discounted Cashflows"
        bar.y_axis.title = "CHF"
        bar.x_axis.title = "Year"
        data = Reference(ws_fin, min_col=2, min_row=1, max_col=4, max_row=len(fin_df) + 1)
        cats = Reference(ws_fin, min_col=1, min_row=2, max_row=len(fin_df) + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws_fin.add_chart(bar, "F2")
        wb.save(out_excel)

    result = {
        "excel_path": str(out_excel),
        "chart_path": str(chart_path),
        "soc_path": str(run_dir / "battery_soc.csv"),
        "financial_summary": financial_results,
    }
    if flow_chart_path is not None:
        result["power_flow_chart_path"] = str(flow_chart_path)
    return result
