import calendar
from logging import config
from pathlib import Path
import pandas as pd
import numpy as np
import tkinter as tk
import re
from tkinter import filedialog
from datetime import datetime, time
from openpyxl import load_workbook
import config

## Will be used to process the input data before feeding the data into the optimization framework

_selected_file_path = None

def get_input_file_path():
    """
    Lazily open file picker once and cache the selected Excel path.
    """
    global _selected_file_path
    if _selected_file_path is None:
        _selected_file_path = filedialog.askopenfilename(
            initialdir="01-INPUT-DATA",
            title="Select Input Data",
            filetypes=[("Excel files", ".xlsx .xls")]
        )
        if not _selected_file_path:
            raise FileNotFoundError("No input data file selected.")
    return _selected_file_path


# ==========================================================
# config reading and generation
# ==========================================================
import math

def write_config_py(config, filename="config.py"):
    with open(filename, "w", encoding="utf-8") as f:
        f.write("# Auto-generated from Excel - DO NOT EDIT MANUALLY\n\n")

        for key, value in config.items():

            # handle NaN (float)
            if isinstance(value, float) and math.isnan(value):
                f.write(f"{key} = None\n")

            # format strings properly
            elif isinstance(value, str):
                f.write(f'{key} = "{value}"\n')

            # everything else (numbers, booleans)
            else:
                f.write(f"{key} = {value}\n")

def clean_value(v):
    if isinstance(v, str):
        v = v.strip()

        if v.lower() == "true":
            return True
        if v.lower() == "false":
            return False

        try:
            if "." in v:
                return float(v)
            return int(v)
        except:
            return v

    return v

def refresh_config_from_excel():
    get_input_file_path()
    config_df = pd.read_excel(_selected_file_path, sheet_name="config")
    config_dict = {
        row["variable"]: clean_value(row["value"])
        for _, row in config_df.iterrows()
    }
    max_timesteps = config_dict.get("max_timesteps")
    om_raw = config_dict.get("operation_and_maintenance")
    if om_raw is None or (isinstance(om_raw, float) and np.isnan(om_raw)):
        year = config_dict.get("year")
        days_in_year = 366 if (year and calendar.isleap(int(year))) else 365
        # Auto-calc: 10 000 CHF/year scaled to the simulation horizon
        config_dict["operation_and_maintenance"] = 10000 * (
            max_timesteps / (24 * 4 * days_in_year) if not np.isnan(max_timesteps) else 1
        )
    write_config_py(config_dict)

refresh_config_from_excel()

# ==========================================================
# Select sheets from excel file
# ==========================================================
def select_sheets(label_text):
    """
    Opens a file dialog with the excel sheets in the selected excel file
    """

    def selection_list():
        selection.clear()
        for var in vars:
            if var.get() == 1:
                selection.append(sheet_list[vars.index(var)])
        root.destroy()

        

    file_path = get_input_file_path()
    df = pd.read_excel(file_path, sheet_name=None)
    sheet_list = [
        sheet for sheet in df.keys()
        if sheet != "config" and not str(sheet).startswith("_xlnm.")
    ]
    selection = []
    root = tk.Tk()
    root.title("Select Data Sheet(s)")
    root.geometry('500x250')

    # Add a label at the top
    label = tk.Label(root, text=label_text, font=("Arial", 12, "bold"))
    label.pack(side='top', pady=5)

    vars = [tk.IntVar() for sheet in sheet_list]
    btns = [tk.Checkbutton(root, text = sheet, variable = vars[sheet_list.index(sheet)],
                        onvalue=1, offvalue=0) for sheet in sheet_list]

    for btn in btns:
        btn.pack(side = 'top')

    sbtn = tk.Button(root, text = 'Apply & Exit', 
                        command = selection_list) 
    sbtn.pack(side = 'top')      

    root.mainloop()

    return(selection)


def _parse_excel_time(value):
    """
    Convert Excel/Pandas/string time values to a Python time object.
    """
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.time()

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    if isinstance(value, (int, float)):
        # Excel stores times as fractions of a day.
        try:
            base = pd.Timestamp("1899-12-30")
            return (base + pd.to_timedelta(float(value), unit="D")).time()
        except Exception:
            return None

    try:
        parsed = pd.to_datetime(str(value), format="%H:%M:%S", errors="coerce")
        return None if pd.isna(parsed) else parsed.time()
    except Exception:
        return None


def _normalize_cell_ref(ref: str) -> str:
    return ref.replace("$", "").upper()


def _evaluate_excel_formula_cell(worksheet, cell_ref: str, cache: dict):
    """
    Evaluate a small subset of Excel formulas needed by the input workbook.
    Supports direct values plus formulas using SUM/MIN/MAX and arithmetic.
    """
    cell_ref = _normalize_cell_ref(cell_ref)
    if cell_ref in cache:
        return cache[cell_ref]

    cell = worksheet[cell_ref]
    value = cell.value

    if value is None:
        cache[cell_ref] = None
        return None

    if isinstance(value, (int, float)):
        cache[cell_ref] = float(value)
        return cache[cell_ref]

    if isinstance(value, str) and value.startswith("="):
        formula = value[1:].strip()

        def cell_value(ref: str):
            return _evaluate_excel_formula_cell(worksheet, ref, cache) or 0.0

        def range_values(start_ref: str, end_ref: str):
            cells = worksheet[f"{_normalize_cell_ref(start_ref)}:{_normalize_cell_ref(end_ref)}"]
            values = []
            for row in cells:
                for item in row:
                    values.append(_evaluate_excel_formula_cell(worksheet, item.coordinate, cache) or 0.0)
            return values

        python_expr = re.sub(
            r"([A-Z]+\d+):([A-Z]+\d+)",
            lambda match: f'RANGE_VALUES("{match.group(1)}", "{match.group(2)}")',
            formula,
        )
        python_expr = re.sub(r"\bSUM\s*\(", "sum(", python_expr, flags=re.IGNORECASE)
        python_expr = re.sub(r"\bMIN\s*\(", "min(", python_expr, flags=re.IGNORECASE)
        python_expr = re.sub(r"\bMAX\s*\(", "max(", python_expr, flags=re.IGNORECASE)
        python_expr = re.sub(
            r"(?<![A-Z0-9_\"'])\b([A-Z]+\d+)\b(?![A-Z0-9_\"'])",
            lambda match: f'CELL("{match.group(1)}")',
            python_expr,
        )
        python_expr = python_expr.replace("^", "**")

        try:
            evaluated = eval(
                python_expr,
                {"__builtins__": {}},
                {"CELL": cell_value, "RANGE_VALUES": range_values, "sum": sum, "min": min, "max": max},
            )
        except Exception as exc:
            raise ValueError(f"Could not evaluate Excel formula '{value}' in cell {cell_ref}") from exc

        cache[cell_ref] = None if evaluated is None else float(evaluated)
        return cache[cell_ref]

    numeric = pd.to_numeric(value, errors="coerce")
    cache[cell_ref] = None if pd.isna(numeric) else float(numeric)
    return cache[cell_ref]


def _load_lkw_template_from_excel(file_path, sheet_name):
    """
    Read the LKW template with openpyxl so derived Excel columns can be
    recomputed even when the workbook has no cached formula results.
    """
    workbook = load_workbook(file_path, data_only=False, read_only=False)
    try:
        worksheet = workbook[sheet_name]
        formula_cache = {}
        rows = []

        for row_idx in range(3, worksheet.max_row + 1):
            time_value = _parse_excel_time(worksheet[f"A{row_idx}"].value)
            lkw_kw = _evaluate_excel_formula_cell(worksheet, f"E{row_idx}", formula_cache)
            if lkw_kw is None:
                total_kwh = _evaluate_excel_formula_cell(worksheet, f"D{row_idx}", formula_cache)
                if total_kwh is not None:
                    lkw_kw = 4 * total_kwh

            rows.append({"row_idx": row_idx, "time": time_value, "lkw_kW": lkw_kw})

        df = pd.DataFrame(rows)

        # Some templates leave the first quarter-hour unlabeled although the
        # derived power formula is present. Recover that midnight slot.
        first_data_row_missing_time = (
            not df.empty
            and df.iloc[0]["row_idx"] == 3
            and pd.isna(df.iloc[0]["time"])
            and pd.notna(df.iloc[0]["lkw_kW"])
        )
        if first_data_row_missing_time:
            df.loc[df["row_idx"] == 3, "time"] = time(0, 0)

        df = df.dropna(subset=["time", "lkw_kW"]).sort_values("time")
        return df.drop(columns=["row_idx"])
    finally:
        workbook.close()


# ==========================================================
#  Control outliers in power data
# ==========================================================

def clean_power_outliers(df, column_name="value", threshold_quantile=0.999):
    """
    Finds values that are completely unrealistic relative to the 99.9th percentile
    and caps them to that maximum value.
    """
    df = df.copy()
    
    # 1. Determine a strict upper threshold (e.g., top 0.1% of all values)
    cutoff = df[column_name].quantile(threshold_quantile)
    
    # Let's say your 99.9th percentile is 400 kW. If a value hits 1202 kW,
    # it's clearly a data glitch. We multiply the cutoff by an allowance factor (e.g., 2.0).
    max_allowed = cutoff * 2.0
    
    # 2. Identify where the data explodes
    outliers = df[column_name] > max_allowed
    
    if outliers.any():
        num_outliers = outliers.sum()
        print(f"⚠️ Found {num_outliers} outlier(s) in {column_name}. Clipping to {max_allowed:.2f} kW")
        
        # 3. Clip the values to your calculated maximum allowance
        df.loc[outliers, column_name] = max_allowed
        
    return df

# ==========================================================
# 10 → 15 minute conversion
# ==========================================================

def _infer_nominal_interval_minutes(index):
    # Accept a Series, Index or array-like of timestamps and ensure we
    # operate on a DatetimeIndex so DatetimeIndex.to_series() is available.
    dt_index = pd.DatetimeIndex(pd.to_datetime(index))

    if dt_index.empty:
        raise ValueError("Cannot infer nominal timestep interval from timestamps")

    # compute differences between consecutive timestamps (in minutes)
    diffs = dt_index.to_series().diff().dropna().dt.total_seconds().div(60)

    if diffs.empty:
        raise ValueError("Cannot infer nominal timestep interval from timestamps")

    mode = diffs.mode()
    interval_minutes = float(mode.iloc[0]) if not mode.empty else float(diffs.median())
    if interval_minutes <= 0:
        raise ValueError("Invalid inferred timestep interval")
    return interval_minutes


def convert_to_15min(df, column_name="power_kW"):
    """
    Safely converts regular interval power data (kW) to 15-minute
    average power data (kW) using time-weighted resampling.
    """
    df = df.copy()
    if "timestamp" in df.columns:
        df = df.set_index("timestamp")

    df.index = pd.to_datetime(df.index)
    df = df.sort_index()
    df = df[[column_name]].dropna()
    df = df[~df.index.duplicated(keep="first")]

    if df.empty:
        return pd.DataFrame(columns=["timestamp", column_name])

    # Use the nominal interval from the input series so the last point can be
    # extended to a full interval and the overlap-weighted 15-min average is
    # computed correctly.
    interval_minutes = _infer_nominal_interval_minutes(df.index)
    interval = pd.Timedelta(minutes=interval_minutes)

    starts = df.index
    ends = starts[1:].to_list() + [starts[-1] + interval]
    values = df[column_name].to_numpy(dtype=float)

    target_start = starts[0].floor("15Min")
    target_end = pd.to_datetime(ends[-1]).ceil("15Min")
    target_index = pd.date_range(
        start=target_start,
        end=target_end - pd.Timedelta(minutes=15),
        freq="15Min"
    )

    weighted_sum = np.zeros(len(target_index), dtype=float)
    total_seconds = np.zeros(len(target_index), dtype=float)

    for interval_start, interval_end, value in zip(starts, ends, values):
        bin_start = interval_start.floor("15Min")
        while bin_start < interval_end:
            bin_end = bin_start + pd.Timedelta(minutes=15)
            overlap_start = max(interval_start, bin_start)
            overlap_end = min(interval_end, bin_end)
            overlap = (overlap_end - overlap_start).total_seconds()
            if overlap > 0:
                idx = int((bin_start - target_start) // pd.Timedelta(minutes=15))
                weighted_sum[idx] += value * overlap
                total_seconds[idx] += overlap
            bin_start = bin_end

    result = np.where(total_seconds > 0, weighted_sum / total_seconds, np.nan)
    return pd.DataFrame({"timestamp": target_index, column_name: result})


# ==========================================================
# Generic trafo loader
# ==========================================================

def load_trafo(sheet_name):
    """
    Load and clean a transformer sheet.
    Supports sheets with the following average units in the data column:
        - avg[W]
        - avg[kW]
        - avg[Wh]
        - avg[kWh]

    Returns:
        DataFrame with columns:
            timestamp (datetime)
            power_kW (float)
    """

    file_path = get_input_file_path()
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=1
    )

    # Updated regex to include 'kW'
    avg_columns = [
        col for col in df.columns
        if isinstance(col, str) and re.search(r"-avg\[(W|kW|Wh|kWh)\]$", col)
    ]
    if not avg_columns:
        raise ValueError(f"No avg column found in sheet {sheet_name}")

    avg_column = avg_columns[0]
    unit_match = re.search(r"-avg\[(?P<unit>W|kW|Wh|kWh)\]$", avg_column)
    unit = unit_match.group("unit") if unit_match else "W"

    df = df[["Zeit", avg_column]]
    df.columns = ["timestamp", "value"]

    # Try to parse common European GUI format explicitly to avoid
    # pandas' fallback-to-dateutil warning. Fall back to dayfirst parsing.
    ts_sample = df["timestamp"].astype(str).dropna().head(20)
    if not ts_sample.empty and ts_sample.str.match(r"\d{1,2}/\d{1,2}/\d{4} \d{2}:\d{2}:\d{2}").all():
        df["timestamp"] = pd.to_datetime(df["timestamp"], format="%d/%m/%Y %H:%M:%S", errors="coerce")
    else:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce", dayfirst=True)

    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["timestamp", "value"]).sort_values("timestamp")
    df = clean_power_outliers(df, column_name="value")
    df = df.drop_duplicates(subset="timestamp", keep="first")

    # Group 1: Direct Power Units
    if unit == "W":
        df["power_kW"] = df["value"] / 1000.0
    elif unit == "kW":
        df["power_kW"] = df["value"]
        
    # Group 2: Accumulated Energy Units (convert to power based on interval duration)
    else:
        df["energy_kWh"] = df["value"] / 1000.0 if unit == "Wh" else df["value"]
        interval_minutes = _infer_nominal_interval_minutes(df["timestamp"])
        df["power_kW"] = df["energy_kWh"] / (interval_minutes / 60.0)

    df = df[["timestamp", "power_kW"]]
    
    # Handled safely now using the time-weighted upsampling/downsampling method!
    df_15 = convert_to_15min(df, "power_kW")
    return df_15


# ==========================================================
# Grid exchange (Trafo1 + Trafo2)
# ==========================================================
def load_grid_exchange(trafo_sheets):
    """
    Returns structured dataframe with:
        timestamp
        trafo1_kW
        trafo2_kW
        grid_exchange_kW
    """

    df = None
    trafo_cols = []

    for i, sheet in enumerate(trafo_sheets, start=1):
        trafo = load_trafo(sheet)

        col_name = f"trafo{i}_kW"
        trafo = trafo.rename(columns={"power_kW": col_name})

        trafo_cols.append(col_name)

        if df is None:
            df = trafo
        else:
            df = df.merge(trafo, on="timestamp", how="outer")

    # total grid exchange
    df["grid_exchange_kW"] = df[trafo_cols].sum(axis=1)

    return df[["timestamp", *trafo_cols, "grid_exchange_kW"]]


# ==========================================================
# EV demand LKW
# ==========================================================
def generate_lkw_profile(file_path=None, year=None, sheet_name=None):
    """
    Generate full-year charging profile for 2025 in 15-minute intervall from single example day data.

    Assumptions:
        - Sheet contains one typical 15-min weekday
        - Charging only Monday–Friday
        - Saturday and Sunday = 0
        - no seasonality
    """

    if file_path is None:
        file_path = get_input_file_path()

    if year is None:
        print("no year detected for charging profile generation")
        return
    
    if sheet_name is None:
        print("no worksheet selected for charging profile generation")
        return

    df = _load_lkw_template_from_excel(file_path, sheet_name)

    if len(df) != 96:
        raise ValueError(
            f"LKW profile in sheet '{sheet_name}' must contain 96 valid 15-min rows after parsing. "
            f"Found {len(df)} rows."
        )

    daily_profile = df["lkw_kW"].values

    # Create full-year 15-min index
    start = pd.Timestamp(f"{year}-01-01 00:00:00")
    end = pd.Timestamp(f"{year}-12-31 23:45:00")
    full_index = pd.date_range(start=start, end=end, freq="15min")

    result = pd.DataFrame({"timestamp": full_index})
    result["weekday"] = result["timestamp"].dt.weekday  # Monday=0, Sunday=6

    # Initialize with 0
    result["lkw_kW"] = 0.0

    # Apply profile only on weekdays
    weekday_mask = result["weekday"] < 5

    # Repeat daily profile for number of weekdays
    weekday_indices = result[weekday_mask].index
    num_weekdays = len(weekday_indices) // 96

    repeated_profile = np.tile(daily_profile, num_weekdays)

    result.loc[weekday_mask, "lkw_kW"] = repeated_profile

    result = result.drop(columns="weekday")

    return result

# ==========================================================
# EV demand Zustellung
# ==========================================================
def generate_zustellung_profile(file_path=None, year=None, sheet_name=None):
    """
    Generate full-year 15-min Zustellung load profile.

    - Winter profile from Excel
    - Mixed resolution (hourly + 15-min) harmonized to 15-min
    - Weekday-dependent (Mon–Sat)
    - Sunday = 0
    - Summer months reduced by 40%
    """

    if file_path is None:
        file_path = get_input_file_path()

    if year is None:
        print("no year detected for delivery profile generation")
        return
    
    if sheet_name is None:
        print("no worksheet selected for delivery profile generation")
        return

    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=0
    )

    # Rename first column to "time" before removing other Unnamed columns
    df = df.rename(columns={df.columns[0]: "time"})

    # Remove empty Excel columns and normalize column names
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
    df.columns = df.columns.astype(str).str.strip()

    # Robust time parsing
    def parse_time(value):
        try:
            if isinstance(value, pd.Timestamp):
                return value.time()
            return pd.to_datetime(value, format="%H:%M:%S").time()
        except:
            return None

    df["time"] = df["time"].apply(parse_time)
    df = df.dropna(subset=["time"])

    # Convert weekday columns to numeric
    weekday_cols = df.columns.drop("time")
    for col in weekday_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.sort_values("time").reset_index(drop=True)

    # harmonize to 15 minute resolution
    df["timestamp"] = (
        pd.to_datetime("1900-01-01")
        + pd.to_timedelta(df["time"].astype(str))
    )

    df = df.set_index("timestamp")
    df = df.drop(columns=["time"])

    full_index = pd.date_range(
        start="1900-01-01 00:00:00",
        end="1900-01-01 23:45:00",
        freq="15min"
    )

    df_15 = df.reindex(full_index).ffill()

    df_15 = df_15.reset_index().rename(columns={"index": "timestamp"})
    df_15["time"] = df_15["timestamp"].dt.time
    df_15 = df_15.drop(columns=["timestamp"])

    # full year timestamp
    start = pd.Timestamp(f"{year}-01-01 00:00:00")
    end = pd.Timestamp(f"{year}-12-31 23:45:00")

    full_year_index = pd.date_range(start=start, end=end, freq="15min")

    result = pd.DataFrame({"timestamp": full_year_index})
    result["weekday_num"] = result["timestamp"].dt.weekday
    result["time"] = result["timestamp"].dt.time

    # mapping weekday profiles
    lookup = {}
    for day in df_15.columns:
        if day != "time":
            lookup[day] = dict(zip(df_15["time"], df_15[day]))

    weekday_map = {
        0: "Montag",
        1: "Dienstag",
        2: "Mittwoch",
        3: "Donnerstag",
        4: "Freitag",
        5: "Samstag"
    }

    result["zustellung_kW"] = 0.0

    for wd_num, wd_name in weekday_map.items():
        mask = result["weekday_num"] == wd_num
        result.loc[mask, "zustellung_kW"] = (
            result.loc[mask, "time"].map(lookup[wd_name])
        )

    # Sunday remains 0 automatically

    # apply summer reduction of -40%
    summer_months = [4, 5, 6, 7, 8, 9]  # April–September

    summer_mask = result["timestamp"].dt.month.isin(summer_months)
    result.loc[summer_mask, "zustellung_kW"] *= 0.6

    result = result[["timestamp", "zustellung_kW"]]

    return result


# ==========================================================
# Energy price curve
# ==========================================================


def load_price_curve(year: int) -> pd.DataFrame:
    """
    Loads energy prices. Handles import and export independently based on config.
    If the spot price file year differs from the load year, remap the spot price
    pattern onto the requested load year.
    """
    # 1. Determine requirements from config
    spot_price_year = getattr(config, "Spot_price_year", year)
    price_csv = Path(f"01-INPUT-DATA/Spot_prices/GUI_ENERGY_PRICES_{spot_price_year}.csv")

    needs_import_dyn = not getattr(config, "use_constant_import_price", False)
    needs_export_dyn = not getattr(config, "use_constant_export_price", False)

    df_dynamic_resampled = None
    if needs_import_dyn or needs_export_dyn:
        if not price_csv.exists():
            raise FileNotFoundError(
                f"Energy price file not found at {price_csv.resolve()}. "
                "Place GUI_ENERGY_PRICES.csv in 01-INPUT-DATA/."
            )

        df = pd.read_csv(price_csv)
        df["timestamp"] = pd.to_datetime(
            df["MTU (CET/CEST)"].str.split(" - ").str[0].str.strip()
            .str.replace(r"\s*\(CET\)|\s*\(CEST\)", "", regex=True),
            format="%d/%m/%Y %H:%M:%S"
        )

        rate = getattr(config, "exchange_rate", 1.0)
        df["electricity_price"] = (
            pd.to_numeric(df["Day-ahead Price (EUR/MWh)"], errors="coerce")
            * rate
            / 1000
        )

        df = df[["timestamp", "electricity_price"]].groupby("timestamp").first()

        full_index = pd.date_range(
            start=df.index.min(),
            end=df.index.max(),
            freq="15min"
        )
        df_dynamic_resampled = df.reindex(full_index).ffill().bfill().reset_index()
        df_dynamic_resampled.columns = ["timestamp", "electricity_price"]

        if spot_price_year != year:
            target_index = pd.date_range(
                start=pd.Timestamp(f"{year}-01-01 00:00:00"),
                end=pd.Timestamp(f"{year}-12-31 23:45:00"),
                freq="15min"
            )
            mapping = df_dynamic_resampled.copy()
            mapping["day_of_year"] = mapping["timestamp"].dt.dayofyear
            mapping["time"] = mapping["timestamp"].dt.time

            target = pd.DataFrame({"timestamp": target_index})
            target["day_of_year"] = target["timestamp"].dt.dayofyear
            target["time"] = target["timestamp"].dt.time

            target = target.merge(
                mapping[["day_of_year", "time", "electricity_price"]],
                on=["day_of_year", "time"],
                how="left"
            )
            target["electricity_price"] = target["electricity_price"].ffill().bfill()
            df_dynamic_resampled = target

    start_date = pd.Timestamp(f"{year}-01-01 00:00:00")
    end_date = pd.Timestamp(f"{year}-12-31 23:45:00")
    final_index = pd.date_range(start=start_date, end=end_date, freq="15min")
    df_prices = pd.DataFrame(index=final_index)

    if needs_import_dyn or needs_export_dyn:
        # Ensure df_dynamic_resampled is aligned to the full year index
        df_dynamic_resampled = df_dynamic_resampled.set_index("timestamp").reindex(final_index).ffill().bfill().reset_index()

    if not needs_import_dyn:
        df_prices["import_price"] = getattr(config, "energy_import_price", 0.0)
    else:
        df_prices["import_price"] = df_dynamic_resampled["electricity_price"].values

    if not needs_export_dyn:
        df_prices["export_price"] = getattr(config, "energy_export_price", 0.0)
    else:
        df_prices["export_price"] = df_dynamic_resampled["electricity_price"].values

    return df_prices.reset_index().rename(columns={"index": "timestamp"})