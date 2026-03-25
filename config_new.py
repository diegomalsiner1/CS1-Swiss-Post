from pathlib import Path

import openpyxl


CONFIG_SHEET_NAME = "CONFIG_INPUTS"
# Optional explicit path to the source Excel file.
# If None, the first .xlsx under 01-INPUT-DATA (excluding TEMPLATES) is used.
SOURCE_EXCEL = None


def _default_parameters() -> dict:
    max_steps = 40000
    return {
        "use_case": "Peak_Shaving",
        "interest_rate": 0.06,
        "lifetime": 20,
        "year": 2025,
        "load_existing_input_dict": False,
        "max_timesteps": max_steps,
        "optimization_mode": "lp",
        "PV_max_capacity": 10000,
        "Battery_max_inflow": 1000,
        "Battery_max_outflow": 1000,
        "Battery_max_capacity": 100000,
        "eta_charge": 0.9,
        "eta_discharge": 0.95,
        "eta_self_discharge": 0.0,
        "invest_cost": 450.0,
        "operation_and_maintenance": 10000 * (max_steps / (24 * 4 * 365)),
        "battery_degrading": 0.01,
        "peak_shaving_cost_factor": 10.0,
        "peak_shaving_granularity": "monthly", ## either "yearly" or "monthly"
    }


def _find_source_excel() -> Path:
    if SOURCE_EXCEL is not None:
        p = Path(SOURCE_EXCEL)
        if not p.is_absolute():
            p = Path.cwd() / p
        if not p.exists():
            raise FileNotFoundError(f"Configured SOURCE_EXCEL does not exist: {p}")
        return p

    input_data_dir = Path.cwd() / "01-INPUT-DATA"
    if not input_data_dir.exists():
        raise FileNotFoundError(f"Input data directory not found: {input_data_dir}")

    candidates = sorted(
        p
        for p in input_data_dir.rglob("*.xlsx")
        if "TEMPLATES" not in p.parts and not p.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(
            f"No .xlsx files found under {input_data_dir}. Put one file in 01-INPUT-DATA first."
        )
    return candidates[0]


def _parse_bool(value, default):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return default
    if isinstance(value, bool):
        return value

    value_str = str(value).strip().lower()
    if value_str in {"true", "1", "yes", "y"}:
        return True
    if value_str in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Cannot parse boolean value: {value}")


def _parse_optional_int(value, default):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return default
    value_str = str(value).strip().lower()
    if value_str == "none":
        return None
    return int(float(value))


def _parse_value(value, default):
    if isinstance(default, bool):
        return _parse_bool(value, default)
    if default is None:
        return _parse_optional_int(value, default)
    if isinstance(default, int):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return default
        return int(float(value))
    if isinstance(default, float):
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return default
        return float(value)
    if isinstance(default, str):
        if value is None:
            return default
        value_str = str(value).strip()
        return value_str if value_str else default
    return value if value is not None else default


def load_runtime_config() -> dict:
    defaults = _default_parameters()
    parse_defaults = dict(defaults)
    parse_defaults["operation_and_maintenance"] = None

    source_excel = _find_source_excel()
    wb = openpyxl.load_workbook(source_excel, data_only=True)

    if CONFIG_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{CONFIG_SHEET_NAME}' not found in {source_excel}. "
            "Create/update it with prepare_input_excel.ipynb first."
        )

    ws = wb[CONFIG_SHEET_NAME]
    excel_values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        param = row[0]
        value = row[1] if len(row) > 1 else None
        if param is None:
            continue
        key = str(param).strip()
        if key:
            excel_values[key] = value

    loaded = {}
    for key, default in parse_defaults.items():
        loaded[key] = _parse_value(excel_values.get(key), default)

    if loaded["operation_and_maintenance"] is None:
        max_steps = loaded["max_timesteps"]
        loaded["operation_and_maintenance"] = 10000 * (
            max_steps / (24 * 4 * 365) if max_steps is not None else 1
        )

    return loaded


def _load_or_default() -> dict:
    try:
        return load_runtime_config()
    except Exception as exc:
        print(f"Warning: Excel config not loaded ({exc}). Falling back to defaults.")
        return _default_parameters()


_cfg = _load_or_default()

use_case = _cfg["use_case"]
interest_rate = _cfg["interest_rate"]
lifetime = _cfg["lifetime"]
year = _cfg["year"]

load_existing_input_dict = _cfg["load_existing_input_dict"]
max_timesteps = _cfg["max_timesteps"]
optimization_mode = _cfg["optimization_mode"]

PV_max_capacity = _cfg["PV_max_capacity"]
Battery_max_inflow = _cfg["Battery_max_inflow"]
Battery_max_outflow = _cfg["Battery_max_outflow"]
Battery_max_capacity = _cfg["Battery_max_capacity"]
eta_charge = _cfg["eta_charge"]
eta_discharge = _cfg["eta_discharge"]
eta_self_discharge = _cfg["eta_self_discharge"]
invest_cost = _cfg["invest_cost"]
operation_and_maintenance = _cfg["operation_and_maintenance"]
battery_degrading = _cfg["battery_degrading"]
peak_shaving_cost_factor = _cfg["peak_shaving_cost_factor"]
peak_shaving_granularity = _cfg["peak_shaving_granularity"]
